from odoo import models, fields, api
import xmlrpc.client
from datetime import datetime
import re
from xml.sax.saxutils import escape
from odoo.exceptions import UserError
import io
import base64
import xlsxwriter
from openpyxl import load_workbook
import logging
_logger = logging.getLogger(__name__)

class DataMigration(models.Model):
    _name = 'data.migration.connection'
    _description = 'Migrate Data from Odoo 13'

    source_url = fields.Char(string="URL", default="http://123.201.19.81:8069")
    source_db = fields.Char(string="Database Name", default="micro_access")
    source_user = fields.Char(string="Database Username", default="admin@microaccess.in")
    source_password = fields.Char(string="Database Password", default="123")

    product_limit = fields.Integer(string="Product limit")
    product_offset = fields.Integer(string="Product offset")

    contact_limit = fields.Integer(string="Contact limit")
    contact_offset = fields.Integer(string="Contact Offset")

    sale_limit = fields.Integer(string="Sale limit")
    sale_offset = fields.Integer(string="Sale offset")

    validate_sale_limit = fields.Integer(string="Validate Sale limit")
    validate_sale_offset = fields.Integer(string="Validate Sale offset")

    purachse_limit = fields.Integer(string="Purachse limit")
    purachse_offset = fields.Integer(string="Purachse offset")

    helpdesk_limit = fields.Integer(string="Helpdesk limit")
    helpdesk_offset = fields.Integer(string="Helpdesk Offset")

    crm_limit = fields.Integer(string="CRM Limit")
    crm_offset = fields.Integer(string="CRM Offset")

    tracking_limit = fields.Integer(string="Tracking Limit")
    tracking_offset = fields.Integer(string="Tracking Offset")
    
    def _connect(self):
        common = xmlrpc.client.ServerProxy(f"{self.source_url}/xmlrpc/2/common")
        uid = common.authenticate(self.source_db, self.source_user, self.source_password, {})
        models = xmlrpc.client.ServerProxy(f"{self.source_url}/xmlrpc/2/object")
        return uid, models

    # def clean_str(self, s):
    #     if not s:
    #         return ''
    #     s = str(s)
    #     s = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', s)
    #     s = s.replace('\n', ' ')  # remove invalid chars
    #     s = s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    #     return s

    def clean_str(self, text):
        """ Cleans text to be safe for XML-RPC migration.
            (This is primarily for cleaning data *after* it's received, 
             but we include robust logic for general use).
        """
        if not isinstance(text, str):
            return text
        
        # 1. Escape basic XML characters: &, <, >
        cleaned_text = escape(text)
        
        # 2. Remove most common XML-invalid control characters 
        # (keeps \t, \n, \r which are usually safe)
        cleaned_text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\uFFFD]', '', cleaned_text)
        
        return cleaned_text

    def safe_clean_data(self, record):
        """Recursively clean all string fields in a record dict."""
        for key, value in record.items():
            if isinstance(value, str):
                record[key] = self.clean_str(value)
            elif isinstance(value, list):
                # If list of tuples/lists, clean recursively
                record[key] = [
                    self.clean_str(v) if isinstance(v, str) else v for v in value
                ]
            elif isinstance(value, dict):
                record[key] = self.safe_clean_data(value)
        return record

    def validate_sale_order_delivery(self):
        uid, models = self._connect()
        start_time = datetime.now()
        # Get all product template IDs from source DB
        _logger.info("this is limit %s", self.validate_sale_limit)
        _logger.info("this is offset %s", self.validate_sale_offset)
        sale_order = models.execute_kw(
            self.source_db, uid, self.source_password,
            'sale.order', 'search_read',
            [[]],
            {
                'fields': [
                    'id', 'name','picking_ids'
                ],
                'limit': self.validate_sale_limit,
                'offset': self.validate_sale_offset,
                'order': 'id asc'  # fetch in ascending order of IDs
            }
        )
        for sale in sale_order:
            _logger.info("this is sale data %s", sale)
            source_record_id = sale['id']
            existing_order = self.env['sale.order'].sudo().search([('source_record_id', '=', source_record_id)], limit=1)
            if existing_order:
                if existing_order.state == 'sale':
                    _logger.info("this is existing order %s", existing_order)
                    if sale['picking_ids']:
                        _logger.info("13 has piking to validate")
                        for picking in sale['picking_ids']:
                            _logger.info("picking %s", picking)
                            stock_picking = models.execute_kw(
                                self.source_db, uid, self.source_password,
                                'stock.picking', 'search_read',
                                [[('id', '=', picking)]],
                                {
                                    'fields': [
                                        'id', 'name','state'
                                    ] 
                                }
                            )
                            # _logger.info('this is stock picking data %s',stock_picking)
                            if stock_picking and stock_picking[0]['state'] == 'done':
                                _logger.info("Stock picking %s is done", stock_picking[0]['name'])
                                current_stock_picking = existing_order.picking_ids
                                current_stock_picking.button_validate()


        end_time = datetime.now()
        self.validate_sale_offset += self.validate_sale_limit
        _logger.info("update offset %s", self.validate_sale_offset)
        _logger.info(f"Total time taken: {end_time - start_time}")

    def migrate_products_tracking(self):
        uid, models = self._connect()
        start_time = datetime.now()

        _logger.info("Migrating tracking fields | offset=%s | limit=%s",
                    self.tracking_offset, self.tracking_limit)

        error_lines = []
        compare_lines = [] 

        # Read only ID + tracking
        products = models.execute_kw(
            self.source_db, uid, self.source_password,
            'product.template', 'search_read',
            [[('active', 'in', [True, False])]],
            {
                'fields': ['id', 'tracking'],
                'limit': self.tracking_limit,
                'offset': self.tracking_offset,
                'order': 'id asc'
            }
        )

        for p in products:
            _logger.info("this is product data %s", p)

            src_id = p.get('id')
            src_tracking = p.get('tracking')

            dest_product = self.env['product.template'].sudo().with_context(active_test=False).search([
                ('source_record_id', '=', src_id)
            ], limit=1)

            _logger.info("Processing source product id %s with tracking %s", src_id, src_tracking)

            if not dest_product:
                msg = f"Destination product not found for source id {src_id}"
                _logger.warning(msg)
                error_lines.append([src_id, "", "", msg])
                continue

            try:

                # ----------------------------
                #   NEW LOGIC
                # ----------------------------

                _logger.info(
                    "COMPARE BEFORE UPDATE | Source ID=%s | Dest ID=%s | Source tracking=%s  | Dest current tracking=%s | Dest current is_storable=%s",
                    src_id,
                    src_tracking,
                    dest_product.id,
                    dest_product.tracking,
                    dest_product.is_storable
                )

                # If source tracking is blank → keep blank
                if not src_tracking:
                    update_vals = {
                        'tracking': False,
                        'is_storable': False   # change to True if needed
                    }

                else:
                    # Normal supported mapping
                    tracking_map = {
                        'serial': ('serial', True),
                        'lot': ('lot', True),
                        'none': ('none', True),
                    }

                    dest_tracking, dest_is_storable = tracking_map.get(
                        src_tracking,
                        (False, False)    # default if unknown
                    )

                    update_vals = {
                        'tracking': dest_tracking,
                        'is_storable': dest_is_storable
                    }

                # Write values
                dest_product.write(update_vals)

                compare_lines.append([
                    src_id,                       
                    dest_product.id,               
                    dest_product.name,             
                    src_tracking,                  
                    dest_product.tracking,        
                    dest_product.is_storable       
                ])

                # Log after update
                _logger.info(
                    "COMPARE AFTER UPDATE Dest ID=%s | Source tracking=%s | NEW tracking=%s | NEW is_storable=%s",
                    dest_product.id,
                    src_tracking,
                    dest_product.tracking,
                    dest_product.is_storable
                )

            except Exception as e:
                error_msg = str(e)
                _logger.error("Error updating product %s: %s", dest_product.id, error_msg)

                error_lines.append([
                    src_id,
                    dest_product.id,
                    dest_product.name,
                    error_msg
                ])
                continue

        end_time = datetime.now()
        self.tracking_offset += self.tracking_limit

        _logger.info("update offset %s", self.tracking_offset)
        _logger.info(f"Total time taken: {end_time - start_time}")

        # Generate Excel file if errors exist
        if error_lines or compare_lines:

            existing_attachment = self.env['ir.attachment'].search([
                ('name', '=', 'tracking_update_errors.xlsx')
            ], limit=1)

            if existing_attachment:
                # Load existing file & append
                file_content = base64.b64decode(existing_attachment.datas)
                existing_file = io.BytesIO(file_content)

                workbook = load_workbook(existing_file)

                # ------------------- ERROR SHEET -------------------
                if 'Tracking Errors' not in workbook.sheetnames:
                    err_sheet = workbook.create_sheet('Tracking Errors')
                    err_sheet.append(["Source ID", "Destination ID", "Product Name", "Error Message"])
                else:
                    err_sheet = workbook['Tracking Errors']

                for line in error_lines:
                    err_sheet.append(line)

                # ------------------- COMPARE SHEET -------------------
                if 'Tracking Compare' not in workbook.sheetnames:
                    cmp_sheet = workbook.create_sheet('Tracking Compare')
                    cmp_sheet.append([
                        "Source Product ID",
                        "Dest Product ID",
                        "Product Name",
                        "Source Tracking",
                        "Dest Tracking (After)",
                        "Dest Is Storable (After)"
                    ])
                else:
                    cmp_sheet = workbook['Tracking Compare']

                for line in compare_lines:
                    cmp_sheet.append(line)

                # Save back to attachment
                output = io.BytesIO()
                workbook.save(output)
                output.seek(0)

                existing_attachment.write({
                    'datas': base64.b64encode(output.read())
                })

                _logger.info("Updated existing Excel file with new rows.")

            else:
                # Create NEW Excel
                output = io.BytesIO()
                workbook = xlsxwriter.Workbook(output, {'in_memory': True})

                # ------------------- ERROR SHEET -------------------
                err_sheet = workbook.add_worksheet('Tracking Errors')
                err_sheet_headers = ["Source ID", "Destination ID", "Product Name", "Error Message"]
                for col, h in enumerate(err_sheet_headers):
                    err_sheet.write(0, col, h)

                row = 1
                for line in error_lines:
                    for col, v in enumerate(line):
                        err_sheet.write(row, col, v)
                    row += 1

                # ------------------- COMPARE SHEET -------------------
                cmp_sheet = workbook.add_worksheet('Tracking Compare')
                cmp_headers = [
                    "Source Product ID",
                    "Dest Product ID",
                    "Product Name",
                    "Source Tracking",
                    "Dest Tracking (After)",
                    "Dest Is Storable (After)"
                ]
                for col, h in enumerate(cmp_headers):
                    cmp_sheet.write(0, col, h)

                row = 1
                for line in compare_lines:
                    for col, v in enumerate(line):
                        cmp_sheet.write(row, col, v)
                    row += 1

                workbook.close()
                output.seek(0)

                # Create attachment
                self.env['ir.attachment'].create({
                    'name': 'tracking_update_errors.xlsx',
                    'type': 'binary',
                    'datas': base64.b64encode(output.read()),
                    'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                })

                _logger.info("Created new Excel file for errors + compare data.")

    def migrate_crm(self):
        uid, models = self._connect()
        start_time = datetime.now()

        _logger.info("CRM migration started - Limit: %s | Offset: %s", self.crm_limit, self.crm_offset)

        leads = models.execute_kw(
            self.source_db, uid, self.source_password,
            'crm.lead', 'search_read',
            [[['type', '=', 'lead']]],
            {
                'fields': [
                    'id', 'name', 'type', 'partner_id', 'partner_name', 'email_from', 'phone', 'mobile',
                    'contact_name', 'street', 'street2', 'city', 'state_id', 'country_id', 'zip', 'function',
                    'user_id', 'team_id', 'stage_id', 'company_id', 'description','title','probability',
                    'priority', 'priority1', 'date_open', 'contact_name2', 'mobile2','website',
                    'tag_ids', 'pipeline_ids', 'call_type', 'date_closed','create_date','lang_id'
                ],
                'limit': self.crm_limit,
                'offset': self.crm_offset,
                'order': 'id asc'
            }
        )

        # _logger.info("Total CRM records fetched: %s", len(leads))

        for lead in leads:
            _logger.info("this is lead data %s",lead)
            source_id = lead.get('id')
            existing_lead = self.env['crm.lead'].sudo().with_context(active_test=False).search([
                ('source_record_id', '=', source_id)
            ], limit=1)

            if existing_lead:
                _logger.info("Skipping existing lead: %s", existing_lead.name)
                continue

            partner_id = False
            if lead.get('partner_id'):
                source_partner_id = lead['partner_id'][0]
                partner = self.env['res.partner'].sudo().with_context(active_test=False).search(
                    [('source_record_id', '=', source_partner_id)], limit=1
                )
                partner_id = partner.id if partner else False

            user_id = False
            if lead.get('user_id'):
                source_user_id = lead['user_id'][0]
                source_user_data = models.execute_kw(
                    self.source_db, uid, self.source_password,
                    'res.users', 'read',
                    [[source_user_id]],
                    {'fields': ['login', 'name']}
                )[0]

                source_login = source_user_data.get('login')
                source_name = source_user_data.get('name')

                user = self.env['res.users'].sudo().with_context(active_test=False).search(
                    [('login', '=', source_login)], limit=1
                )
                if not user:
                    user = self.env['res.users'].sudo().with_context(active_test=False).search(
                        [('name', '=', source_name)], limit=1
                    )
                if user:
                    user_id = user.id

            team_id = False
            if lead.get('team_id'):
                team_name = self.clean_str(lead['team_id'][1])
                team = self.env['crm.team'].search([('name', '=', team_name)], limit=1)
                if not team:
                    team = self.env['crm.team'].create({'name': team_name})
                team_id = team.id

            stage_id = False
            if lead.get('stage_id'):
                stage_name = self.clean_str(lead['stage_id'][1])
                stage = self.env['crm.stage'].search([('name', '=', stage_name)], limit=1)
                if not stage:
                    stage = self.env['crm.stage'].create({'name': stage_name})
                stage_id = stage.id

            country_id = False
            if lead.get('country_id'):
                country_name = self.clean_str(lead['country_id'][1])
                country = self.env['res.country'].search([('name', '=', country_name)], limit=1)
                country_id = country.id if country else False

            state_id = False
            if lead.get('state_id'):
                state_name = self.clean_str(lead['state_id'][0])
                state = self.env['res.country.state'].search([('id', '=', state_name)], limit=1)
                state_id = state.id

            title_id = False
            if lead.get('title'):
                title_name = self.clean_str(lead['title'][1])
                title = self.env['res.partner.title'].search([('name','=',title_name)],limit=1)
                title_id = title.id

            lang_id = False
            if lead.get('lang_id'):
                lang_name = self.clean_str(lead['lang_id'][1])
                lang = self.env['res.lang'].search([('name','=',lang_name)],limit=1)
                lang_id = lang.id

            tag_ids = []
            if lead.get('tag_ids'):
                for tag_id in lead['tag_ids']:
                    tag_data = models.execute_kw(
                        self.source_db, uid, self.source_password,
                        'crm.lead.tag', 'read', [[tag_id]], {'fields': ['name']}
                    )[0]
                    tag_name = self.clean_str(tag_data.get('name'))
                    tag = self.env['crm.tag'].search([('name', '=', tag_name)], limit=1)
                    if not tag:
                        tag = self.env['crm.tag'].create({'name': tag_name})
                    tag_ids.append(tag.id)

            pipeline_ids = []
            if lead.get('pipeline_ids'):
                for pl_id in lead['pipeline_ids']:
                    pipeline_data = models.execute_kw(
                        self.source_db, uid, self.source_password,
                        'pipeline.master', 'read', [[pl_id]], {'fields': ['pipeline']}
                    )[0]
                    pl_name = self.clean_str(pipeline_data.get('pipeline'))
                    pipeline = self.env['pipeline.master'].search([('pipeline', '=', pl_name)], limit=1)
                    if not pipeline:
                        pipeline = self.env['pipeline.master'].create({'pipeline': pl_name})
                    pipeline_ids.append(pipeline.id)

            priority_value = lead.get('priority1')

            if priority_value in ['Cold', '0', 0]:
                priority_value = '0'
            elif priority_value in ['Warm', '1', 1]:
                priority_value = '1'
            elif priority_value in ['Hot', '2', 2]:
                priority_value = '2'
            else:
                priority_value = '0'

            name = lead.get('name')
            if not name:
                # Use partner name or a fallback
                partner_name = lead.get('partner_name') or "Unknown Partner"
                name = f"{partner_name} (Migrated Lead ID {lead.get('id')})"
                _logger.warning("⚠️ Lead ID %s has no name — using fallback: %s", lead.get('id'), name)

            record = self.env['crm.lead'].sudo().create({
                'source_record_id': source_id,
                'name': name,
                'type': lead.get('type', 'lead'),
                'partner_id': partner_id,
                'partner_name': self.clean_str(lead.get('partner_name')),
                'email_from': lead.get('email_from'),
                'phone': lead.get('phone'),
                'mobile': lead.get('mobile'),
                'contact_name': self.clean_str(lead.get('contact_name')),
                'contact_name2': self.clean_str(lead.get('contact_name2')),
                'mobile2': lead.get('mobile2'),
                'function': self.clean_str(lead.get('function')),
                'street': self.clean_str(lead.get('street')),
                'street2': self.clean_str(lead.get('street2')),
                'city': self.clean_str(lead.get('city')),
                'zip': lead.get('zip'),
                'state_id': state_id,
                'country_id': country_id,
                'user_id': user_id,
                'team_id': team_id,
                'stage_id': stage_id,
                'priority': lead.get('priority'),
                'priority1': priority_value,
                'call_type': lead.get('call_type'),
                'description': lead.get('description'),
                'date_open': lead.get('date_open'),
                'date_closed': lead.get('date_closed'),
                'tag_ids': [(6, 0, tag_ids)],
                'pipeline_ids': [(6, 0, pipeline_ids)],
                'title': title_id,
                'probability': lead.get('probability'),
                'website': lead.get('website'),
                'language' : lang_id
            })

            src_date = lead.get('create_date')

            # ✅ Only update if source has a real date
            if src_date:
                try:
                    # Handle both date and datetime formats
                    if len(src_date) == 10:
                        formatted_date = datetime.strptime(src_date, '%Y-%m-%d')
                    else:
                        formatted_date = datetime.strptime(src_date, '%Y-%m-%d %H:%M:%S')

                    formatted_date_str = formatted_date.strftime('%Y-%m-%d %H:%M:%S')

                    # ✅ Update create_date & write_date only if valid source date
                    self.env.cr.execute("""
                        UPDATE crm_lead
                        SET create_date = %s,
                            write_date = %s
                        WHERE id = %s
                    """, (formatted_date_str, formatted_date_str, record.id))

                    self.env.cr.commit()
                    _logger.info("✅ Lead ID %s create_date set to %s", record.id, formatted_date_str)

                except Exception as e:
                    _logger.error("❌ Error updating create_date for lead %s: %s", record.id, e)

            else:
                # ✅ Do nothing (keep Odoo’s default create_date)
                _logger.info("ℹ️ No date in source for lead ID %s — keeping default create_date", record.id)

        end_time = datetime.now()
        _logger.info("end_time : %s", end_time)
        self.crm_offset += self.crm_limit
        _logger.info("update offset %s", self.crm_offset)
        _logger.info(f"Total time taken: {end_time - start_time}")


    def migrate_sales(self):
        uid, models = self._connect()
        start_time = datetime.now()
        # Get all product template IDs from source DB
        _logger.info("this is limit %s", self.sale_limit)
        _logger.info("this is offset %s", self.sale_offset)
        sale_order = models.execute_kw(
            self.source_db, uid, self.source_password,
            'sale.order', 'search_read',
            [[]],
            {
                'fields': [
                    'id', 'name', 'partner_id', 'customer_status', 'subject', 'partner_shipping_contact', 'partner_shipping_mobile','user_id',
                    'date_order', 'payment_term_id', 'customer_po_no', 'customer_po_date', 'sale_quotation', 'expected_delivery_date', 'order_line', 'revision_ids', 'state','cancel_remarks','remarks_cancels_id'
                ],
                'limit': self.sale_limit,
                'offset': self.sale_offset,
                'order': 'id asc'  # fetch in ascending order of IDs
            }
        )
        for sale in sale_order:
            _logger.info("this is sale data %s", sale)
            sale_ref_id = sale['id']
            existing_order = self.env['sale.order'].sudo().search([('source_record_id', '=', sale_ref_id)], limit=1)
            if not existing_order:
                order_line_ids = sale.get('order_line', [])
                order_line_vals = []
                history_data = []

                if order_line_ids:
                    order_line_data = models.execute_kw(
                        self.source_db, uid, self.source_password,
                        'sale.order.line', 'read',
                        [order_line_ids],
                        {
                            'fields': [
                                'id', 'order_id', 'product_id', 'name',
                                'product_uom_qty', 'price_unit',
                                'tax_id', 'price_subtotal', 'display_type'
                            ]
                        }
                    )

                    for order_line in order_line_data:
                        _logger.info("this is order line %s", order_line)

                        line_name = self.clean_str(order_line.get('name'))

                        if not order_line['display_type'] in ['line_section', 'line_note']:
                            product = self.env['product.template'].sudo().with_context(active_test=False).search([('source_record_id', '=', order_line['product_id'][0])], limit=1)
                            if product:
                                product_id = product.id
                            # _logger.info("this is product id %s", product)

                            product_id = product.product_variant_id.id  # ensure correct product.product
                            # _logger.info("this is product variant id %s", product_id)

                            # product_uom_id = False
                            # if order_line.get('product_uom'):
                            #     uom_name = order_line['product_uom'][1]  # name of UoM from source
                            #     uom = self.env['uom.uom'].search([('name', '=', uom_name)], limit=1)
                            #     if uom:
                            #         product_uom_id = uom.id

                            tax_ids = []
                            tax_record = self.env['account.tax'].search([('id', '=', 26)], limit=1)
                            if tax_record:
                                tax_ids.append(tax_record.id)

                            order_line_vals.append((0, 0, {
                                'source_record_id':order_line['id'],
                                'product_id': product_id,
                                'name': line_name,
                                # 'product_uom': product_uom_id,
                                'product_uom_qty': order_line.get('product_uom_qty', 0.0),
                                'price_unit': order_line.get('price_unit', 0.0),
                                'tax_id': [(6, 0, tax_ids)],
                            }))
                        elif order_line['display_type'] in ['line_section', 'line_note']:
                            order_line_vals.append((0, 0, {
                                'source_record_id':order_line['id'],
                                # 'product_id': False,
                                'name': line_name,
                                'display_type':order_line['display_type']
                            }))


                    # revision / history logic (unchanged)
                    
                    if sale.get('revision_ids'):
                        for revision_id in sale['revision_ids']:
                            revisions = models.execute_kw(
                                self.source_db, uid, self.source_password,
                                'revision.history', 'read',
                                [[revision_id]],
                                {'fields': ['id', 'name', 'revision_date', 'sale_id', 'history_line_ids']}
                            )
                            _logger.info("this is revisions %s",revisions)
                            if revisions:
                                revision = revisions[0]
                                history_lines = []

                                for history_line_id in revision.get('history_line_ids', []):
                                    history_line = models.execute_kw(
                                        self.source_db, uid, self.source_password,
                                        'revision.history.line', 'read',
                                        [[history_line_id]],
                                        {'fields': ['id', 'revision_history_ids', 'description', 'unit_price', 'product_id', 'qty']}
                                    )[0]
                                    _logger.info("this is history lines %s",history_line)

                                    # Find mapped product
                                    product_id_source = history_line.get('product_id')
                                    product_id = False
                                    if product_id_source and isinstance(product_id_source, list):
                                        product = self.env['product.template'].sudo().with_context(active_test=False).search(
                                            [('source_record_id', '=', product_id_source[0])], limit=1
                                        )
                                        if product:
                                            product_id = product.product_variant_id.id

                                    history_line_desc = self.clean_str(order_line.get('description'))

                                    history_lines.append((0, 0, {
                                        'source_record_id': history_line['id'],
                                        'revision_history_ids': history_line.get('revision_history_ids', []),
                                        'description': history_line_desc,
                                        'unit_price': history_line.get('unit_price'),
                                        'product_id': product_id,
                                        'qty': history_line.get('qty'),
                                    }))

                                history_data.append((0, 0, {
                                    'source_record_id': revision['id'],
                                    'name': revision.get('name'),
                                    'revision_date': revision.get('revision_date'),
                                    'history_line_ids': history_lines,
                                }))

                # partner mapping (unchanged)
                partner_id = False
                if sale.get('partner_id'):
                    _logger.info("this is fetch partenr id %s", sale.get('partner_id'))
                    partner = self.env['res.partner'].sudo().with_context(active_test=False).search([('source_record_id', '=', sale['partner_id'][0])], limit=1)
                    if partner:
                        partner_id = partner.id
                # _logger.info("this is partner %s", partner_id)

                partner_invoice_id = False
                partner_shipping_id = False
                if sale.get('partner_invoice_id'):
                    partner_invoice = self.env['res.partner'].search([('source_record_id', '=', sale['partner_invoice_id'][0])], limit=1)
                    if partner_invoice:
                        partner_invoice_id = partner_invoice.id

                if sale.get('partner_shipping_id'):
                    partner_ship = self.env['res.partner'].search([('source_record_id', '=', sale['partner_shipping_id'][0])], limit=1)
                    if partner_ship:
                        partner_shipping_id = partner_ship.id

                # Payment term, currency, company, team
                payment_term_id = False
                if sale.get('payment_term_id'):
                    payment_term_name = sale['payment_term_id'][1]
                    payment_term = self.env['account.payment.term'].search([('name', '=', payment_term_name)], limit=1)
                    if not payment_term:
                        payment_term = self.env['account.payment.term'].create({'name': payment_term_name})
                    payment_term_id = payment_term.id

                terms_id = False
                if sale.get('sale_quotation'):
                    term_name = sale['sale_quotation']
                    if term_name == 'othersale':
                        term_name = 'Other Sale'
                    elif term_name == 'amcsale':
                        term_name = 'AMC Sale'
                    terms = self.env['terms.conditions'].search([('name', '=', term_name)], limit=1)
                    if not terms:
                        terms = self.env['terms.conditions'].create({'name': term_name})
                    terms_id = terms.id

                cancel_remarks_id = False
                if sale.get('cancel_remarks'):
                    remarks = sale['cancel_remarks']
                    remark = self.env['remarks.remarks'].search([('remarks', '=', remarks[1])], limit=1)
                    if not remark:
                        remark = self.env['remarks.remarks'].create({'remarks': remarks[1], 'source_record_id': remarks[0]})
                    cancel_remarks_id = remark.id

                remarks_cancel_id = False
                if sale.get('remarks_cancels_id'):
                    cancel_remarks = sale['remarks_cancels_id']
                    remarks_cancel_id = self.env['crm.lost.reason'].search([('name', '=', cancel_remarks[1])], limit=1)
                    _logger.info("cancel remark %s", remarks_cancel_id)
                    if not remarks_cancel_id:
                        _logger.info("not found cancel remark %s", cancel_remarks[1])
                        remarks_cancel_id = self.env['crm.lost.reason'].create({'name': cancel_remarks[1]})
                        _logger.info("creating cancel remark %s", remarks_cancel_id)
                    remarks_cancel_id = remarks_cancel_id.id

                # Create customer_status
                customer_status = 'newcustomer'
                if sale.get('customer_status'):
                    if sale.get('customer_status') == 'daler':
                        customer_status = 'dealer'
                    if sale.get('customer_status') == 'existingcustomer':
                        customer_status = 'existingcustomer'
                    if sale.get('customer_status') == 'newcustomer':
                        customer_status = 'newcustomer'

                user_id = False
                if sale.get('user_id'):
                    source_user_id = sale['user_id'][0]
                    source_user_data = models.execute_kw(
                        self.source_db, uid, self.source_password,
                        'res.users', 'read',
                        [[source_user_id]],
                        {'fields': ['login', 'name']}
                    )[0]

                    source_login = source_user_data.get('login')
                    source_name = source_user_data.get('name')

                    # Try to find user in destination by login first, then name
                    user = self.env['res.users'].sudo().search([('login', '=', source_login)], limit=1)
                    if not user:
                        user = self.env['res.users'].sudo().search([('name', '=', source_name)], limit=1)

                    if user:
                        user_id = user.id

                # Create the sale order with its lines
                new_order = self.env['sale.order'].sudo().create({
                    'source_record_id': sale_ref_id,
                    'name': sale.get('name'),
                    'partner_id': partner_id,
                    'user_id': user_id,
                    'customer_status': customer_status,
                    'subject': sale.get('subject'),
                    'partner_shipping_contact': sale.get('partner_shipping_contact'),
                    'partner_shipping_mobile': sale.get('partner_shipping_mobile'),
                    'date_order': sale.get('date_order'),
                    'payment_term_id': payment_term_id,
                    'client_order_ref': sale.get('customer_po_no'),
                    'customer_po_date': sale.get('customer_po_date'),
                    # 'terms_id': terms_id,
                    'validity_date': sale.get('expected_delivery_date'),
                    'order_line': order_line_vals,
                    'revision_ids': history_data,
                    'cancel_remarks': cancel_remarks_id,
                    'remarks_cancel_ids': remarks_cancel_id
                })

                _logger.info("Created sale order %s with %d lines", new_order.name, len(order_line_vals))
                if new_order:
                    new_order.terms_id = terms_id
                    ######################### Rikta: New code ######################################
                    if terms_id:
                        terms = self.env['terms.conditions'].browse(terms_id)
                        if terms and terms.description:
                            new_order.note = terms.description
                    ########################### New code ends #####################################
                    if sale['state'] == "cancel":
                        new_order.action_cancel()
                    elif sale['state'] == 'sale':
                        new_order.action_confirm()
                     # ❌ remove Odoo’s auto-created waiting pickings
                    for auto_picking in new_order.picking_ids:
                        auto_picking.action_cancel()
                        auto_picking.unlink()
                   

                # ---------------- AUTO STOCK TRANSFER (fixed: single picking per order) ----------------
                # Build list of lines that have qty_delivered in source DB
                delivered_lines = []  # list of dicts: {'line_rec': <sale.order.line>, 'qty_delivered': float, 'qty_invoiced': float}
                for so_line in new_order.order_line:
                    if not so_line.product_id:
                        # skip notes/sections
                        continue

                    # Fetch qty delivered from Odoo 13 via XMLRPC for this line
                    try:
                        line_data_13 = models.execute_kw(
                            self.source_db, uid, self.source_password,
                            'sale.order.line', 'read',
                            [[so_line.source_record_id]],
                            {'fields': ['qty_delivered', 'qty_invoiced']}
                        )[0]
                    except Exception as e:
                        _logger.exception("Failed to fetch source line data for source id %s: %s", so_line.source_record_id, e)
                        line_data_13 = {}

                    qty_delivered_13 = line_data_13.get('qty_delivered', 0.0)
                    qty_invoiced_13 = line_data_13.get('qty_invoiced', 0.0)

                    # If qty_delivered_13 is None or 0.0, skip; otherwise record it
                    try:
                        qty_delivered_13 = float(qty_delivered_13 or 0.0)
                    except Exception:
                        qty_delivered_13 = 0.0
                    try:
                        qty_invoiced_13 = float(qty_invoiced_13 or 0.0)
                    except Exception:
                        qty_invoiced_13 = 0.0

                    if qty_delivered_13 > 0:
                        delivered_lines.append({
                            'so_line': so_line,
                            'qty_delivered': qty_delivered_13,
                            'qty_invoiced': qty_invoiced_13,
                        })
                    else:
                        # still update qty_invoiced if present in source
                        if qty_invoiced_13 > 0:
                            try:
                                so_line.sudo().write({'qty_invoiced': qty_invoiced_13})
                            except Exception:
                                _logger.exception("Unable to write qty_invoiced on line %s", so_line.id)

                # If any delivered lines present, create a single picking and moves
                picking = False
                if delivered_lines:
                    picking_type = self.env['stock.picking.type'].search([('code', '=', 'outgoing')], limit=1)
                    if not picking_type:
                        _logger.warning("No outgoing picking type found; skipping auto-delivery for order %s", new_order.name)
                    else:
                        # create one picking for the order
                        try:
                            picking = self.env['stock.picking'].sudo().create({
                                'partner_id': new_order.partner_id.id or False,
                                'picking_type_id': picking_type.id,
                                'origin': new_order.name,
                                'sale_id': new_order.id,
                                'location_id': picking_type.default_location_src_id.id or False,
                                'location_dest_id': picking_type.default_location_dest_id.id or False,
                            })
                        except Exception as e:
                            _logger.exception("Failed to create picking for order %s: %s", new_order.name, e)
                            picking = False

                        # create moves & move lines for each delivered line
                        for dl in delivered_lines:
                            so_line = dl['so_line']
                            qty_to_ship = dl['qty_delivered']
                            qty_invoiced_13 = dl['qty_invoiced']

                            # pick a product_uom: use line.product_uom if present, otherwise product.uom_id
                            product_uom_id = so_line.product_uom.id if so_line.product_uom else (so_line.product_id.uom_id.id if so_line.product_id and so_line.product_id.uom_id else False)
                            if not product_uom_id:
                                # fallback to 1.0 uom if very unusual case
                                uom = self.env['uom.uom'].search([], limit=1)
                                product_uom_id = uom.id if uom else False

                            try:
                                move = self.env['stock.move'].sudo().create({
                                    'name': so_line.name or so_line.product_id.display_name,
                                    'product_id': so_line.product_id.id,
                                    'product_uom_qty': qty_to_ship,
                                    'product_uom': product_uom_id,
                                    'picking_id': picking.id,
                                    'location_id': picking_type.default_location_src_id.id,
                                    'location_dest_id': picking_type.default_location_dest_id.id,
                                    'sale_line_id': so_line.id,
                                })
                            except Exception as e:
                                _logger.exception("Failed to create stock.move for line %s: %s", so_line.id, e)
                                continue

                            # create stock.move.line for done quantity (force done)
                            try:
                                self.env['stock.move.line'].sudo().create({
                                    'move_id': move.id,
                                    'product_id': so_line.product_id.id,
                                    'product_uom_id': product_uom_id,
                                    'qty_done': qty_to_ship,
                                    'location_id': picking_type.default_location_src_id.id,
                                    'location_dest_id': picking_type.default_location_dest_id.id,
                                })
                            except Exception as e:
                                _logger.exception("Failed to create stock.move.line for move %s: %s", move.id, e)

                            # update qty_invoiced on the sale.order.line (so that invoiced qty is preserved)
                            try:
                                so_line.sudo().write({'qty_invoiced': qty_invoiced_13})
                            except Exception:
                                _logger.exception("Unable to write qty_invoiced on line %s after creating moves", so_line.id)

                        # Confirm and validate the picking once
                        if picking:
                            try:
                                picking.action_confirm()
                                # Auto-validate if picking is ready
                                if picking.state in ['assigned', 'confirmed']:
                                    picking.button_validate()
                            except Exception as e:
                                _logger.exception("Failed to validate picking %s", picking.id)

                            # After validation, explicitly update sale order lines' delivered quantities if needed
                            for dl in delivered_lines:
                                so_line = dl['so_line']
                                shipped_qty = dl['qty_delivered']
                                try:
                                    # Some versions of Odoo compute qty_delivered; writing may fail if computed.
                                    # We try to write it; if it fails, at least stock moves exist which is the source of truth.
                                    so_line.sudo().write({'qty_delivered': shipped_qty})
                                except Exception:
                                    # If write fails (computed field), log and continue
                                    _logger.debug("Could not write qty_delivered on sale.order.line %s; it may be computed. shipped %s", so_line.id, shipped_qty)

                            _logger.info("Stock picking validated for order %s: applied delivered quantities", new_order.name)

        end_time = datetime.now()
        self.sale_offset += self.sale_limit
        _logger.info("update offset %s", self.sale_offset)
        _logger.info(f"Total time taken: {end_time - start_time}")

    def migrate_purchases(self):
        uid, models = self._connect()
        start_time = datetime.now()

        _logger.info("this is limit %s", self.purachse_limit)
        _logger.info("this is offset %s", self.purachse_offset)

        # Fetch Purchase Orders
        purchase_orders = models.execute_kw(
            self.source_db, uid, self.source_password,
            'purchase.order', 'search_read',
            [[]],
            {
                'fields': [
                    'id', 'name', 'partner_id', 'partner_ref','date_order','against_selection', 'shipping_address', 'purchase_sale_order2', 'currency_id',
                    'order_line', 'state', 'amount_total', 'notes', 'date_order','date_approve','payment_term_id','user_id'
                ],
                'limit': self.purachse_limit,
                'offset': self.purachse_offset,
                'order': 'id asc'
            }
        )

        for po in purchase_orders:
            _logger.info("this is purchase order data %s", po)
            po_ref_id = po['id']

            # Skip already migrated POs
            existing_po = self.env['purchase.order'].sudo().search([('source_record_id', '=', po_ref_id)], limit=1)
            if existing_po:
                _logger.info("PO %s already migrated. Skipping.", po.get('name'))
                continue

            order_line_ids = po.get('order_line', [])
            order_line_vals = []

            if order_line_ids:
                order_line_data = models.execute_kw(
                    self.source_db, uid, self.source_password,
                    'purchase.order.line', 'read',
                    [order_line_ids],
                    {'fields': [
                        'id', 'order_id', 'product_id', 'name',
                        'product_qty', 'price_unit', 'taxes_id',
                        'qty_received', 'qty_invoiced','display_type'
                    ]}
                )

                for line in order_line_data:
                    _logger.info("this is purchase order line %s", line)

                    line_name = self.clean_str(line.get('name'))
                    
                    if not line['display_type'] in ['line_section', 'line_note']:
                        product = self.env['product.template'].sudo().with_context(active_test=False).search(
                            [('source_record_id', '=', line['product_id'][0])], limit=1)
                        if product:
                            product_id = product.id
                            # _logger.info("this is product id %s", product)
                        product_id = product.product_variant_id.id
                        # _logger.info("this is product variant id %s", product_id)

                        # Taxes
                        tax_ids = []
                        tax_record = self.env['account.tax'].search([('id', '=', 57)], limit=1)
                        if tax_record:
                            tax_ids.append(tax_record.id)

                        order_line_vals.append((0, 0, {
                            'source_record_id': line['id'],
                            'product_id': product_id,
                            'name': line_name,
                            'product_qty': line.get('product_qty', 0.0),
                            'price_unit': line.get('price_unit', 0.0),
                            'taxes_id': [(6, 0, tax_ids)],
                            'qty_invoiced': line.get('qty_invoiced', 0.0),
                        }))

                    elif line['display_type'] in ['line_section', 'line_note']:
                            order_line_vals.append((0, 0, {
                                'source_record_id':line['id'],
                                'name': line_name,
                                'display_type':line['display_type'],
                                'product_qty': 0.0,
                                'price_unit': 0.0,
                            }))
            # Partner
            partner_id = False
            if po.get('partner_id'):
                partner = self.env['res.partner'].sudo().with_context(active_test=False).search(
                    [('source_record_id', '=', po['partner_id'][0])], limit=1)
                if partner:
                    partner_id = partner.id
            
            payment_term_id = False 
            if po.get('payment_term_id'): 
                term_name = po['payment_term_id'][1] 
                payment_term = self.env['account.payment.term'].search([('name', '=', term_name)], limit=1) 
                if not payment_term: 
                    payment_term = self.env['account.payment.term'].create({'name': term_name}) 
                payment_term_id = payment_term.id

            shipping_address = False
            source_shipping = po.get('shipping_address')

            if source_shipping:
                if source_shipping in ['Micro Access', 'micro']:
                    shipping_address = 'micro'
                elif source_shipping in ['Order', 'order']:
                    shipping_address = 'order'
                else:
                    shipping_address = 'micro'
            else:
                # Fallback if no value found
                shipping_address = 'micro'
            
            # Map related sale orders 
            sale_order_ids = [] 
            if po.get('purchase_sale_order2'): 
                for so_id in po['purchase_sale_order2']: 
                    sale_order = self.env['sale.order'].sudo().with_context(active_test=False).search( [('source_record_id', '=', so_id)], limit=1 ) 
                    if sale_order: 
                        sale_order_ids.append(sale_order.id)

            user_id = False
            if po.get('user_id'):
                source_user_id = po['user_id'][0]
                source_user_data = models.execute_kw(
                    self.source_db, uid, self.source_password,
                    'res.users', 'read',
                    [[source_user_id]],
                    {'fields': ['login', 'name']}
                )[0]

                source_login = source_user_data.get('login')
                source_name = source_user_data.get('name')

                # Try to find user in destination by login first, then name
                user = self.env['res.users'].sudo().with_context(active_test=False).search([('login', '=', source_login)], limit=1)
                if not user:
                    user = self.env['res.users'].sudo().with_context(active_test=False).search([('name', '=', source_name)], limit=1)

                if user:
                    user_id = user.id

            # Create PO
            new_po = self.env['purchase.order'].sudo().create({
                'source_record_id': po_ref_id,
                'name': po.get('name'),
                'partner_id': partner_id,
                # 'currency_id': currency_id,
                'date_order': po.get('date_order') or fields.Datetime.now(),
                'date_approve': po.get('date_approve'),
                'against_selection': po.get('against_selection'), 
                'shipping_address': shipping_address, 
                'purchase_sale_order2': [(6, 0, sale_order_ids)],
                'payment_term_id': payment_term_id, 
                'partner_ref': po.get('partner_ref'),
                'notes': po.get('notes'),
                'order_line': order_line_vals,
                'user_id':user_id
            })

            _logger.info("Created purchase order %s with %d lines", new_po.name, len(order_line_vals))

            # Set PO state
            if po['state'] == 'cancel':
                new_po.button_cancel()
            elif po['state'] == 'purchase':
                new_po.button_confirm()

            # Remove auto-created waiting receipts
            for auto_picking in new_po.picking_ids:
                try:
                    if auto_picking.state not in ['done', 'cancel']:
                        auto_picking.action_cancel()
                        auto_picking.unlink()
                        _logger.info("🗑️ Removed Odoo auto-picking %s for PO %s", auto_picking.name, new_po.name)
                except Exception as e:
                    _logger.warning("Could not remove picking for PO %s: %s", new_po.name, e)

            # ---------------- AUTO STOCK RECEIPT (like sales picking) ----------------
            received_lines = []

            for line in new_po.order_line:
                try:
                    source_line_data = models.execute_kw(
                        self.source_db, uid, self.source_password,
                        'purchase.order.line', 'read',
                        [[line.source_record_id]],
                        {'fields': ['qty_received', 'qty_invoiced']}
                    )[0]
                except Exception as e:
                    _logger.warning("Could not read source line %s: %s", line.source_record_id, e)
                    continue

                qty_received_13 = float(source_line_data.get('qty_received') or 0.0)
                qty_invoiced_13 = float(source_line_data.get('qty_invoiced') or 0.0)

                if qty_received_13 > 0:
                    received_lines.append({
                        'line': line,
                        'qty_received': qty_received_13,
                        'qty_invoiced': qty_invoiced_13
                    })

            if received_lines:
                picking_type = self.env['stock.picking.type'].search([('code', '=', 'incoming')], limit=1)
                if not picking_type:
                    _logger.warning("No incoming picking type found; skipping receipts for PO %s", new_po.name)
                else:
                    picking = self.env['stock.picking'].sudo().create({
                        'partner_id': new_po.partner_id.id,
                        'picking_type_id': picking_type.id,
                        'origin': new_po.name,
                        'purchase_id': new_po.id,
                        'location_id': picking_type.default_location_src_id.id,
                        'location_dest_id': picking_type.default_location_dest_id.id,
                    })

                    for rec in received_lines:
                        line = rec['line']
                        qty_received = rec['qty_received']
                        qty_invoiced = rec['qty_invoiced']

                        uom_id = line.product_uom.id or line.product_id.uom_id.id

                        move = self.env['stock.move'].sudo().create({
                            'name': line.name,
                            'product_id': line.product_id.id,
                            'product_uom_qty': qty_received,
                            'product_uom': uom_id,
                            'picking_id': picking.id,
                            'location_id': picking_type.default_location_src_id.id,
                            'location_dest_id': picking_type.default_location_dest_id.id,
                            'purchase_line_id': line.id,
                        })

                        self.env['stock.move.line'].sudo().create({
                            'move_id': move.id,
                            'product_id': line.product_id.id,
                            'product_uom_id': uom_id,
                            'qty_done': qty_received,
                            'location_id': picking_type.default_location_src_id.id,
                            'location_dest_id': picking_type.default_location_dest_id.id,
                        })

                        line.sudo().write({'qty_received': qty_received})

                    try:
                        picking.action_confirm()
                        picking.action_assign()
                        picking.button_validate()
                        _logger.info("Auto receipt validated for PO %s", new_po.name)
                    except Exception as e:
                        _logger.warning("Failed to validate picking for PO %s: %s", new_po.name, e)

            # ------------------- Populate qty_invoiced via vendor bills -------------------
            invoice_line_vals = []
            for line in new_po.order_line:
                try:
                    source_line_data = models.execute_kw(
                        self.source_db, uid, self.source_password,
                        'purchase.order.line', 'read',
                        [[line.source_record_id]],
                        {'fields': ['qty_invoiced', 'taxes_id']}
                    )[0]
                except Exception as e:
                    _logger.warning("Could not read source line %s: %s", line.source_record_id, e)
                    continue

                qty_invoiced_13 = float(source_line_data.get('qty_invoiced') or 0.0)
                if qty_invoiced_13 <= 0:
                    continue

                tax_ids = line.taxes_id.ids or []
                if not tax_ids and source_line_data.get('taxes_id'):
                    tax_ids = source_line_data['taxes_id']

                invoice_line_vals.append((0, 0, {
                    'purchase_line_id': line.id,
                    'product_id': line.product_id.id,
                    'quantity': qty_invoiced_13,
                    'price_unit': line.price_unit,
                    'tax_ids': [(6, 0, tax_ids)],
                }))
            # Only create invoice if we have at least one line
            if invoice_line_vals:
                invoice_vals = {
                    'move_type': 'in_invoice',
                    'partner_id': new_po.partner_id.id,
                    'purchase_id': new_po.id,
                    'invoice_origin': new_po.name,
                    'invoice_date': fields.Date.today(),
                    'invoice_line_ids': invoice_line_vals,
                }

                # Create single invoice
                invoice = self.env['account.move'].sudo().create(invoice_vals)

                try:
                    invoice.action_post()
                    _logger.info("Invoice created for PO %s with %s lines", new_po.name, len(invoice_line_vals))
                except Exception as e:
                    _logger.warning("Could not post invoice for PO line %s: %s", line.id, e)

        end_time = datetime.now()
        self.purachse_offset += self.purachse_limit
        _logger.info("update offset %s", self.purachse_offset)
        _logger.info(f"Total time taken: {end_time - start_time}")

    def migrate_helpdesk(self):
        uid, models = self._connect()
        start_time = datetime.now()
        _logger.info("this is limit %s", self.helpdesk_limit)
        _logger.info("this is offset %s", self.helpdesk_offset)

        # Get only Open tickets
        helpdesk_tickets = models.execute_kw(
            self.source_db, uid, self.source_password,
            'helpdesk.ticket', 'search_read',
            [[('stage_id.name', 'not in', ['Solved', 'Cancelled'])]],
            {
                'fields': [
                    'id', 'name', 'partner_id','user_id','tag_ids','ticket_type_service', 'ticket_type_id',
                    'ticekt_remarks_receving', 'ticket_nature_problem', 'ticket_solve_remarks',
                    'serial_no', 'qty', 'create_date', 'ticket_close_date',
                    'in_progress_date', 'process_solved_date', 'sale_order_id',
                    'current_status', 'note_inward', 'is_repeat_ticket',
                    'service_product_id', 'vendor_id', 'outward_challan_id',
                    'repair_id', 'description', 'stage_id',
                    'assign_history_ids', 'hold_history_ids', 'product_line_ids'
                ],
                'limit': self.helpdesk_limit,
                'offset': self.helpdesk_offset,
                'order': 'id asc'
            }
        )

        for ticket in helpdesk_tickets:
            _logger.info("this is ticket data %s",ticket)
            ticket_ref_id = ticket['id']

            # Skip if already migrated
            if self.env['helpdesk.ticket'].sudo().search([('source_record_id', '=', ticket_ref_id)], limit=1):
                continue

            # Partner
            partner_id = False
            if ticket.get('partner_id'):
                partner = self.env['res.partner'].sudo().with_context(active_test=False).search(
                    [('source_record_id', '=', ticket['partner_id'][0])], limit=1)
                if partner:
                    partner_id = partner.id

            # Sale Order
            sale_order_id = False
            if ticket.get('sale_order_id'):
                sale = self.env['sale.order'].sudo().search(
                    [('source_record_id', '=', ticket['sale_order_id'][0])], limit=1)
                if sale:
                    sale_order_id = sale.id

            # Ticket Type
            ticket_type_id = False
            if ticket.get('ticket_type_id'):
                name = ticket['ticket_type_id'][1]
                ticket_type = self.env['ticket.type'].sudo().search([('name', '=', name)], limit=1)
                if not ticket_type:
                    ticket_type = self.env['ticket.type'].sudo().create({'name': name})
                ticket_type_id = ticket_type.id
                # _logger.info("this is ticket type id %s",ticket_type_id)

            TICKET_SERVICE_MAP = {
                'warranty':'warranty',
                'amc':'amc',
                'chargeable': 'chargeable',           
                'non_chargeable': 'non-chargeable', 
                'fms':'fms',
                'newinstallation':'newinstallation'
            }

            ticket_type_service = TICKET_SERVICE_MAP.get(ticket.get('ticket_type_service'))

            tag_ids = []
            if ticket.get('tag_ids'):
                for t_id in ticket['tag_ids']:
                    tag_data = models.execute_kw(self.source_db, uid, self.source_password,
                                                'helpdesk.tag', 'read', [t_id], {'fields': ['name']})
                    if tag_data:
                        tag_name = tag_data[0]['name']
                        tag = self.env['helpdesk.tag'].sudo().search([('name','=',tag_name)], limit=1)
                        if not tag:
                            tag = self.env['helpdesk.tag'].sudo().create({'name': tag_name})
                        tag_ids.append(tag.id)
            # _logger.info("Mapped tag IDs %s", tag_ids)

            STAGE_NAME_MAP = {
                'Hold': 'On Hold',
                'In Progress': 'In Progress',   
                'New': 'New',                  
                'Solved': 'Solved',
                'Cancelled': 'Cancelled',
            }

            stage_id = False
            if ticket.get('stage_id'):
                source_stage_name = ticket['stage_id'][1]
                # Apply name mapping if necessary
                target_stage_name = STAGE_NAME_MAP.get(source_stage_name, source_stage_name)

                stage = self.env['helpdesk.stage'].sudo().search([('name', '=', target_stage_name)], limit=1)
                if not stage:
                    stage = self.env['helpdesk.stage'].sudo().create({'name': target_stage_name})
                stage_id = stage.id
            # _logger.info("This is current stage name %s",stage_id)

            # M2M Fields
            def map_m2m(source_model, ids, field, dest_model=None):
                res_ids = []
                if ids:
                    # Read from source DB
                    data = models.execute_kw(
                        self.source_db, uid, self.source_password,
                        source_model, 'read', [ids],
                        {'fields': [field]}
                    )
                    for rec in data:
                        # Use destination model for local ORM
                        local_model = dest_model or source_model
                        local = self.env[local_model].sudo().search([(field, '=', rec[field])], limit=1)
                        if not local:
                            local = self.env[local_model].sudo().create({field: rec[field]})
                        res_ids.append(local.id)
                return [(6, 0, res_ids)]

            # model_name = models.execute_kw(
            #     self.source_db, uid, self.source_password,
            #     'ir.model', 'search_read',
            #     [[('model', 'ilike', 'remarks')]],
            #     {'fields': ['model', 'name']}
            # )
            # _logger.info("this is model name %s",model_name)

            ticket_remarks_receiving = map_m2m('remarks.receiving.master', ticket.get('ticket_remarks_receiving'), 'receiving_name','remarks.master')
            ticket_nature_problem = map_m2m('nature.of.problem',ticket.get('ticket_nature_problem'),'nature_problem','nature.problem')
            ticket_solve_remarks = map_m2m('solve.remarks', ticket.get('ticket_solve_remarks'), 'solve_remarks')

            # Users
            user_ids = []
            if ticket.get('user_id'):
                user_source_id = ticket['user_id'][0]
                user_data = models.execute_kw(self.source_db, uid, self.source_password, 'res.users', 'read',
                                            [user_source_id], {'fields': ['name']})
                if user_data:
                    user_name = user_data[0]['name']
                    user = self.env['res.users'].sudo().search([('name', '=', user_name)], limit=1)
                    if user:
                        user_ids.append(user.id)

            # Assign History
            assign_history_vals = []
            if ticket.get('assign_history_ids'):
                assign_data = models.execute_kw(
                    self.source_db, uid, self.source_password, 'assign.history.line', 'read',
                    [ticket['assign_history_ids']],
                    {'fields': ['id', 'assign_date', 'assign_close_date', 'user_id']}
                )
                for rec in assign_data:
                    _logger.info("this is assign history data %s",rec)
                    users = []
                    user_data = rec.get('user_id')
                    if user_data:
                        if isinstance(user_data, list):
                            user = self.env['res.users'].sudo().search([('name', '=', user_data[1])], limit=1)
                            if user:
                                users.append(user.id)
                        elif isinstance(user_data, int):
                            users.append(user_data)
                    assign_history_vals.append((0, 0, {
                        'source_record_id': rec['id'],
                        'assign_date': rec.get('assign_date'),
                        'assign_close_date': rec.get('assign_close_date'),
                        'assigned_to': [(6, 0, users)],
                    }))

            # Hold History
            hold_history_vals = []
            if ticket.get('hold_history_ids'):
                hold_data = models.execute_kw(
                    self.source_db, uid, self.source_password, 'hold.history.line', 'read',
                    [ticket['hold_history_ids']],
                    {'fields': ['id', 'hold_date', 'hold_close_date', 'total_time', 'total_days', 'hold_note', 'hold_closed_note']}
                )
                for rec in hold_data:
                    _logger.info("this is hold history data %s",rec)
                    hold_history_vals.append((0, 0, {
                        'source_record_id': rec['id'],
                        'hold_date': rec.get('hold_date'),
                        'hold_close_date': rec.get('hold_close_date'),
                        'total_time': rec.get('total_time'),
                        'total_days': rec.get('total_days'),
                        'hold_note': rec.get('hold_note'),
                        'hold_closed_note': rec.get('hold_closed_note'),
                    }))

            # Product Lines
            product_line_vals = []
            if ticket.get('product_line_ids'):
                line_data = models.execute_kw(
                    self.source_db, uid, self.source_password, 'line.product', 'read',
                    [ticket['product_line_ids']], {'fields': ['id', 'service_product_id', 'serial_numer', 'quantity']}
                )
                for rec in line_data:
                    _logger.info("this is product line data %s",rec)
                    product_id = False
                    if rec.get('service_product_id'):
                        prod = self.env['product.product'].sudo().with_context(active_test=False).search(
                            [('source_record_id', '=', rec['service_product_id'][0])], limit=1)
                        if prod:
                            product_id = prod.id
                    product_line_vals.append((0, 0, {
                        'source_record_id': rec['id'],
                        'service_product_id': product_id,
                        'serial_numer': rec.get('serial_numer'),
                        'quantity': rec.get('quantity', 0.0),
                    }))

            service_product_id = False
            if ticket.get('service_product_id'):
                source_id, source_name = ticket['service_product_id']

                prod = self.env['product.product'].sudo().with_context(active_test=False).search(
                    [('source_record_id', '=', source_id)],
                    limit=1
                )

                if prod:
                    service_product_id = prod.id
                else:
                    _logger.warning("No matching product found for source_id=%s, source_name=%s. Creating new product.", source_id, source_name)
                    prod = self.env['product.product'].sudo().create({
                        'name': source_name,
                        'source_record_id': source_id,
                        'type': 'consu',  
                        'sale_ok': True,
                        'purchase_ok': True,
                    })
                    service_product_id = prod.id
                    _logger.info("Created new product in destination: id=%s, name=%s", prod.id, prod.name)

            # Create Ticket
            new_ticket = self.env['helpdesk.ticket'].sudo().with_context(skip_assign_history=True).create({
                'source_record_id': ticket_ref_id,
                'name': ticket.get('name'),
                'partner_id': partner_id,
                'tag_ids': [(6, 0, tag_ids)],
                'ticket_type_service': ticket_type_service,
                'ticket_type_id': ticket_type_id,
                'sale_order_id': sale_order_id,
                # 'return_repair_id': repair_order_id,
                # 'outward_challan_id': challan_id,
                'stage_id': stage_id,
                'user_ids': user_ids,
                'ticket_remarks_receiving': ticket_remarks_receiving,
                'ticket_nature_problem': ticket_nature_problem,
                'ticket_solve_remarks': ticket_solve_remarks,
                'serial_no': ticket.get('serial_no'),
                'qty': ticket.get('qty'),
                'create_date': ticket.get('create_date'),
                'ticket_close_date': ticket.get('ticket_close_date'),
                'in_progress_date': ticket.get('in_progress_date'),
                'process_solved_date': ticket.get('process_solved_date'),
                'current_status': ticket.get('current_status'),
                'note_inward': ticket.get('note_inward'),
                'is_repeat_ticket': ticket.get('is_repeat_ticket', False),
                'feedback': ticket.get('feedback'),
                'service_product_id': service_product_id,
                'assign_history_ids': assign_history_vals,
                'hold_history_ids': hold_history_vals,
                'product_line_ids': product_line_vals,
                'description': ticket.get('description'),
            })

            # 🔧 Force update create_date via SQL
            if ticket.get('create_date'):
                query = """
                    UPDATE helpdesk_ticket
                    SET create_date = %s
                    WHERE id = %s
                """
                self.env.cr.execute(query, (ticket['create_date'], new_ticket.id))

            _logger.info(
                "Created Helpdesk Ticket %s (%s) | "
                "Sale Order: %s",
                new_ticket.name,
                new_ticket.id,
                sale_order_id or 'None',
            )

            # 🧩 Repair Order Migration (Inline)
            if ticket.get('repair_id'):
                source_repair_id = ticket['repair_id'][0]

                # Check if repair already exists in destination
                repair = self.env['repair.order'].sudo().search(
                    [('source_record_id', '=', source_repair_id)], limit=1)

                if not repair:
                    # Read from source database (correct source model)
                    repair_data = models.execute_kw(
                        self.source_db, uid, self.source_password,
                        'repair.order.custom', 'read', [source_repair_id],
                        {'fields': [
                            'name', 'partner_id', 'product_id', 'responsible_user_id',
                            'warranty_expiration_date', 'ticket_id', 'subject',
                            'remarks_ids', 'date', 'tag_ids', 'repair_lines_custom_ids',
                            'invoice_method', 'remarks_description','state'
                        ]}
                    )
                    _logger.info("this is repair data %s",repair_data)
                    if repair_data:
                        data = repair_data[0]

                        # Partner mapping
                        partner_id_repair = False
                        if data.get('partner_id'):
                            partner_rec = self.env['res.partner'].sudo().search(
                                [('source_record_id', '=', data['partner_id'][0])], limit=1)
                            partner_id_repair = partner_rec.id if partner_rec else False

                        # Product mapping
                        product_id = False
                        if data.get('product_id'):
                            product = self.env['product.product'].sudo().search(
                                [('source_record_id', '=', data['product_id'][0])], limit=1)
                            product_id = product.id if product else False

                        # Responsible user mapping
                        responsible_user_id = False
                        if data.get('responsible_user_id'):
                            user = self.env['res.users'].sudo().search(
                                [('source_record_id', '=', data['responsible_user_id'][0])], limit=1)
                            responsible_user_id = user.id if user else False

                        # Lines mapping
                        line_vals = []
                        for line_id in data.get('repair_lines_custom_ids', []):
                            line_data = models.execute_kw(
                                self.source_db, uid, self.source_password,
                                'repair.lines.custom', 'read', [[line_id]],
                                {'fields': ['product_id','lot_id','quantity','uom_id','name']}
                            )
                            for l in line_data:
                                _logger.info("this is repair order line %s",l)

                                prod_id = False
                                if l.get('product_id'):
                                    prod = self.env['product.product'].sudo().search(
                                        [('source_record_id', '=', l['product_id'][0])], limit=1)
                                    if prod:
                                        prod_id = prod.id
                                    else:
                                        fallback_prod = self.env['product.product'].sudo().search([('name', 'ilike', 'Service Charges')], limit=1)
                                        prod_id = fallback_prod.id if fallback_prod else False

                                uom_id = self.env['product.product'].browse(prod_id).uom_id.id if prod_id else False

                                line_vals.append((0, 0, {
                                    'name': l.get('name'),
                                    'description_picking': l.get('name'),
                                    'product_id': prod_id,
                                    'product_uom': uom_id,
                                    'product_uom_qty': l.get('quantity') or 1.0,
                                    'company_id': self.env.company.id,
                                    'repair_line_type': 'add',
                                    # 'state': 'draft',
                                }))

                                # _logger.info("Repair order line values: %s", line_vals)

                        # Create in destination database with all extra fields
                        repair = self.env['repair.order'].sudo().create({
                            'source_record_id': source_repair_id,
                            'name': data.get('name'),
                            'partner_id': partner_id_repair,
                            'product_id': product_id,
                            'remarks_description': data.get('remarks_description'),
                            'schedule_date': data.get('date'),
                            'user_ids': responsible_user_id,
                            'warranty_expiration_date': data.get('warranty_expiration_date'),
                            'ticket_id': new_ticket.id,
                            'subject': data.get('subject'),
                            'invoice_method': data.get('invoice_method'),
                            'move_ids': line_vals,
                            'tag_ids': [(6, 0, data.get('tag_ids', []))] if data.get('tag_ids') else False,
                            'state': data.get('state')
                        })

                new_ticket.sudo().write({'return_repair_id': repair.id})
                _logger.info(
                    "✅ Linked Repair Order [ID: %s] to Helpdesk Ticket [ID: %s | Name: %s]",
                    repair.id,
                    new_ticket.id,
                    new_ticket.name
                )

            # 🧩 Challan Migration (Inline)
            if ticket.get('outward_challan_id'):
                source_challan_id = ticket['outward_challan_id'][0]
                
                # Check if challan already exists in destination
                challan = self.env['returnable.goods'].sudo().search(
                    [('source_record_id', '=', source_challan_id)], limit=1)
                
                if not challan:
                    # Read from source
                    challan_data = models.execute_kw(
                        self.source_db, uid, self.source_password,
                        'returnable.goods', 'read', [[source_challan_id]],
                        {'fields': [
                            'challan_no', 'partner_id', 'ticket_id', 'repair_order_custom_id',
                            'return_date', 'customer_id','remarks', 'challan_date', 'returnable_goods_line_ids','state',
                            'return_received_qty_ids','over_estimate','customer_estimate'
                        ]}
                    )
                    _logger.info("this is challan data %s",challan_data)
                    if challan_data:
                        data = challan_data[0]

                        # Partner mapping
                        partner_challan = False
                        if data.get('partner_id'):
                            partner = self.env['res.partner'].sudo().search(
                                [('source_record_id', '=', data['partner_id'][0])], limit=1)
                            partner_challan = partner.id if partner else False
                        _logger.info("partner id %s",partner_challan)

                        # Customer mapping
                        customer_id = False
                        if data.get('customer_id'):
                            customer = self.env['res.partner'].sudo().search(
                                [('source_record_id', '=', data['customer_id'][0])], limit=1)
                            customer_id = customer.id if customer else False
                        _logger.info("this is customer name %s",customer_id)

                        # Repair order mapping
                        repair_order_id = False
                        if data.get('repair_order_custom_id'):
                            repair_order_src_id = data['repair_order_custom_id'][0]
                            repair = self.env['repair.order'].sudo().search(
                                [('source_record_id', '=', repair_order_src_id)], limit=1)
                            repair_order_id = repair.id if repair else False

                        micro_remarks_ids = []
                        if data.get('remarks'):
                            for remark_id in data['remarks']: 
                                micro_remark = self.env['micro.remarks'].sudo().search([('source_record_id', '=', remark_id)], limit=1)
                                if not micro_remark:
                                    remark_data = models.execute_kw(
                                        self.source_db, uid, self.source_password,
                                        'micro.remarks', 'read', [[remark_id]],
                                        {'fields': ['name']}
                                    )
                                    if remark_data:
                                        micro_remark = self.env['micro.remarks'].sudo().create({
                                            'source_record_id': remark_id,
                                            'name': remark_data[0]['name']
                                        })
                                if micro_remark:
                                    micro_remarks_ids.append(micro_remark.id)

                        # Line items
                        line_vals = []
                        for line_id in data.get('returnable_goods_line_ids'):
                            line_data = models.execute_kw(
                                self.source_db, uid, self.source_password,
                                'returnable.goods.line', 'read', [[line_id]],
                                {'fields': ['product_name', 'description', 'serial_no', 'dummy_qty_available','qty', 'return_qty']}
                            )
                            for l in line_data:
                                _logger.info("this is returnable goods data %s",l)
                                
                                line_vals.append((0, 0, {
                                    'product_name': l.get('product_name'),
                                    'description': l.get('description'),
                                    'serial_no': l.get('serial_no'),
                                    'qty_available': l.get('dummy_qty_available'),
                                    'qty': l.get('qty'),
                                    'return_qty': l.get('return_qty'),
                                }))

                        returned_received_vals = []
                        for rr_id in data.get('return_received_qty_ids', []):
                            rr_data = models.execute_kw(
                                self.source_db, uid, self.source_password,
                                'return.received.quantity', 'read', [[rr_id]],
                                {'fields': ['product_name', 'serial_no', 'qty', 'subsidiary_challan_no']}
                            )
                            for rr in rr_data:
                                returned_received_vals.append((0, 0, {
                                    'product_name': rr.get('product_name'),
                                    'serial_no': rr.get('serial_no'),
                                    'qty': rr.get('qty'),
                                    'subsidiary_challan_no': rr.get('subsidiary_challan_no'),
                                }))

                        # Create challan in destination
                        challan = self.env['returnable.goods'].sudo().create({
                            'source_record_id': source_challan_id,
                            'challan_no': data.get('challan_no'),
                            'ticket_id': new_ticket.id,
                            'partner_id': partner_challan,
                            'customer_id': customer_id,
                            'repair_order_id': repair_order_id,  # map to migrated repair order
                            'return_date': data.get('return_date'),
                            'remarks': micro_remarks_ids,
                            'challan_date': data.get('challan_date'),
                            'state': data.get('state'),
                            'over_estimate': data.get('over_estimate'),
                            'customer_estimate': data.get('customer_estimate'),
                            'returnable_goods_line_ids': line_vals,
                            'return_received_qty_ids': returned_received_vals
                        })

                new_ticket.sudo().write({'outward_challan_id': challan.id})
                _logger.info(
                    "✅ Linked Challan [ID: %s] to Helpdesk Ticket [ID: %s | Name: %s]",
                    challan.id,
                    new_ticket.id,
                    new_ticket.name
                )

        end_time = datetime.now()
        _logger.info("end_time : %s", end_time)
        self.helpdesk_offset += self.helpdesk_limit
        _logger.info("update offset %s", self.helpdesk_offset)
        _logger.info(f"Total time taken: {end_time - start_time}")

    def migrate_users(self):
        uid, models = self._connect()
        
        # Get all users from source
        source_users = models.execute_kw(
            self.source_db, uid, self.source_password,
            'res.users', 'search_read',
            [[]],
            {
                'fields': ['id', 'name', 'login', 'partner_id'],
                'order': 'id asc'
            }
        )
        
        migrated_count = 0
        for user in source_users:
            _logger.info("this is user data %s",user)
            # Skip if already migrated
            if self.env['res.users'].sudo().with_context(active_test=False).search([('source_record_id', '=', user['id'])], limit=1):
                continue
            
            # Find the partner that was already migrated
            partner_id = False
            if user.get('partner_id'):
                partner = self.env['res.partner'].sudo().with_context(active_test=False).search(
                    [('source_record_id', '=', user['partner_id'][0])], limit=1)
                if partner:
                    partner_id = partner.id
                else:
                    _logger.warning("Partner not found for user %s. Partner source ID: %s", user['name'], user['partner_id'][0])
            else:
                raise UserError(f"User {user['name']} has no partner_id. Migration halted.")

            # Create user linked to existing partner
            user_vals = {
                # 'source_record_id': user['id'],
                'name': user.get('name'),
                'login': user.get('login'),
                'email': user.get('email', ''),
                'partner_id': partner_id,
            }
            
            new_user = self.env['res.users'].sudo().create(user_vals)
            migrated_count += 1
            _logger.info("Migrated user: %s (ID: %s) -> %s (ID: %s)", 
                        user['name'], user['id'], new_user.name, new_user.id)
        
        _logger.info("Total users migrated: %s", migrated_count)
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'User Migration Complete',
                'message': f'Successfully migrated {migrated_count} users',
                'type': 'success',
                'sticky': False,
            }
        }

    def migrate_products(self):
        uid, models = self._connect()
        
        start_time = datetime.now()
        # Get all product template IDs from source DB
        _logger.info("this is offset %s", self.product_offset)
        _logger.info("this is limit %s", self.product_limit)
        products = models.execute_kw(
            self.source_db, uid, self.source_password,
            'product.template', 'search_read',
            [[('active', 'in', [True, False])]],
            {
                'fields': [
                    'id', 'name', 'sale_ok', 'purchase_ok', 'type',
                    'categ_id', 'list_price', 'standard_price', 'product_brand_id', 'active', 'uom_id', 'uom_po_id', 'l10n_in_hsn_code'
                ],
                'limit': self.product_limit,
                'offset': self.product_offset,
                'order': 'id asc'  # fetch in ascending order of IDs
            }
        )
        
        for p in products:
            _logger.info("product is %s", p)
            product_ref_id = p.get('id')
            _logger.info("this is product ref id %s", product_ref_id)
            existing_product = self.env['product.template'].sudo().with_context(active_test=False).search([('source_record_id', '=', product_ref_id)])
            if not existing_product:
                _logger.info("this is new product create")
                name = self.clean_str(p.get('name')) or "Unnamed Product"

                # 🔹 Map Category by name (create if missing)
                categ_id = False
                if p.get('categ_id'):
                    # categ_name = p['categ_id'][1]  # (id, name)
                    categ_name = self.clean_str(p['categ_id'][1]) if p.get('categ_id') else ''
                    category = self.env['product.category'].search([('name', '=', categ_name)], limit=1)
                    if not category:
                        category = self.env['product.category'].create({'name': categ_name})
                    categ_id = category.id

                brand_id = False
                if p.get('product_brand_id'):
                    # brand_name = p['product_brand_id'][1]  # (id, name)
                    brand_name = self.clean_str(p['product_brand_id'][1]) if p.get('product_brand_id') else ''
                    brand = self.env['product.brand'].search([('brand_name', '=', brand_name)], limit=1)
                    if not brand:
                        brand = self.env['product.brand'].create({'brand_name': brand_name})
                    brand_id = brand.id
                
                product_type = 'consu'
                is_storeable = True
                if p.get('type'):
                    p_type = p.get('type')
                    if p_type == 'consu':
                        product_type = 'consu'
                        is_storeable = False
                    elif p_type == "service":
                        product_type = 'service'
                        is_storeable = False
                    elif p_type == 'product':
                        product_type = 'consu'
                        is_storeable = True

                uom_id = False
                if p.get('uom_id'):
                    uom_name = self.clean_str(p['uom_id'][1]) if p.get('uom_id') else ''
                    uom = self.env['uom.uom'].search([('name', '=', uom_name)], limit=1)
                    if not uom:
                        uom = self.env['uom.uom'].create({'name': uom_name})
                    uom_id = uom.id

                uom_po_id = False
                if p.get('uom_po_id'):
                    uom_name = self.clean_str(p['uom_po_id'][1]) if p.get('uom_po_id') else ''
                    uom_po = self.env['uom.uom'].search([('name', '=', uom_name)], limit=1)
                    if not uom_po:
                        uom_po = self.env['uom.uom'].create({'name': uom_name})
                    uom_po_id = uom_po.id
                        
                # 🔹 Create product in Odoo 18
                self.env['product.template'].create({
                    'source_record_id': product_ref_id,
                    'name': name,
                    'sale_ok': p.get('sale_ok', True),
                    'purchase_ok': p.get('purchase_ok', True),
                    'type': product_type,
                    'is_storable' : is_storeable,
                    'categ_id': categ_id,
                    'brand_id': brand_id,
                    'list_price': p.get('list_price') or 0.0,
                    'standard_price': p.get('standard_price') or 0.0,
                    'uom_id': uom_id,
                    'uom_po_id': uom_po_id,
                    'l10n_in_hsn_code': p.get('l10n_in_hsn_code', ''),
                    'active' : p.get('active')
                })
            
        end_time = datetime.now()
        _logger.info("end_time : %s", end_time)
        self.product_offset += self.product_limit
        _logger.info("update offset %s", self.product_offset)
        _logger.info(f"Total time taken: {end_time - start_time}")

    def migrate_contacts(self):
        uid, models = self._connect()
        
        start_time = datetime.now()
        _logger.info("this is contact limit %s", self.contact_limit)
        _logger.info("this is contact offset %s", self.contact_offset)

        contacts = models.execute_kw(
            self.source_db, uid, self.source_password,
            'res.partner', 'search_read',
            [[('active', 'in', [True, False])]],
            {
                'fields': [
                    'id', 'name', 'phone', 'mobile', 'email',
                    'city', 'state_id', 'country_id', 'vat', 'active', 'street', 'street2', 'zip','contact_person','customer_support_email','category_master_id','user_id','partner_debit', 'partner_credit'
                ],
                'limit': self.contact_limit,
                'offset': self.contact_offset,
                'order': 'id asc'  # fetch in ascending order of IDs
            }
        )
        _logger.info("contact %s",contacts)

        for c in contacts:
            _logger.info("this is contact %s", c)
            contact_ref_id = c.get('id')
            existing_contact = self.env['res.partner'].sudo().with_context(active_test=False).search([('source_record_id', '=', contact_ref_id)])
            _logger.info("Existing %s",existing_contact)
            if not existing_contact:
                _logger.info("create new contact")

                country_id = False
                if c.get('country_id'):
                    country_name = self.clean_str(c['country_id'][1]) if c.get('country_id') else ''
                    country = self.env['res.country'].search([('name', '=', country_name)], limit=1)
                    country_id = country.id
                    
                    # if not country:
                    #     country = self.env['res.country'].create({'name': country_name})
                    # country_id = country.id

                state_id = False
                if c.get('state_id'):
                    state_name = self.clean_str(c['state_id'][1]) if c.get('state_id') else ''
                    state = self.env['res.country.state'].search([('name', '=', state_name)], limit=1)
                    # if not state:
                    #     state = self.env['res.country.state'].create({'name': state_name})
                    state_id = state.id

                category_ids = False
                if c.get('category_master_id'):
                    # Get name from source data
                    cat_name = c['category_master_id'][1]
                    category = self.env['category.master'].search([('name', '=', cat_name[1])], limit=1)
                    if not category:
                        category = self.env['category.master'].create({'name': cat_name})
                    category_ids = category.id


                user_id = False
                if c.get('user_id'):
                    source_user_id = c['user_id'][0]
                    target_user = self.env['res.users'].search([('source_record_id', '=', source_user_id)], limit=1)
                    if target_user:
                        user_id = target_user.id
                

                self.env['res.partner'].create({
                    'source_record_id': contact_ref_id,
                    'name': self.clean_str(c.get('name')) or "Unnamed Contact",
                    'phone': c.get('phone'),
                    'mobile': c.get('mobile'),
                    'email': c.get('email'),
                    'city': self.clean_str(c.get('city')),
                    'state_id': state_id,
                    'country_id': country_id or False,
                    'vat': c.get('vat'),
                    'street': self.clean_str(c.get('street')),
                    'street2': self.clean_str(c.get('street2')),
                    'zip': c.get('zip'),
                    'active': c.get('active', True),
                    'company_type': c.get('company_type', 'person'),
                    'contact_person': c.get('contact_person'),
                    'customer_support_email': c.get('customer_support_email'),
                    'category_ids': category_ids,
                    'user_id': user_id,
                    # 'category_id': [(6, 0, tag_ids)],
                    'debit': c.get('partner_debit'),
                    'credit': c.get('partner_credit'),
                })

        end_time = datetime.now()
        _logger.info("end_time : %s", end_time)
        self.contact_offset += self.contact_limit
        _logger.info("update offset %s", self.contact_offset)
        _logger.info(f"Total time taken: {end_time - start_time}")
